import openpyxl


def rows_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def read_data(file, sheetname, rowno,colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row = rowno, column=colno).value

def read_write(file, sheetname, rowno,colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row = rowno, column=colno).value = data
    workbook.save(file)





